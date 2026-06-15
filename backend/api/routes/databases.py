import os
from io import BytesIO
from typing import Dict, Optional

import polars as pl
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse

from core.database_manager import DatabaseError, DatabaseManager
from .files import _load_database_into_session
from .. import session

router = APIRouter()
db_manager = DatabaseManager()


@router.get("")
async def list_databases() -> Dict:
    databases = db_manager.list_databases()
    return {
        "databases": [d.to_dict() for d in databases],
        "count": len(databases),
    }


@router.get("/current")
async def get_current_database() -> Dict:
    session_id = "default"
    name = session.current_database.get(session_id) or ""
    if not name:
        return {"database": None}
    return {"database": name}


@router.get("/{name}")
async def open_database(name: str) -> Dict:
    session_id = "default"
    if not db_manager.database_exists(name):
        raise HTTPException(status_code=404, detail=f"Database '{name}' not found")
    try:
        result = _load_database_into_session(session_id, name)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error opening database: {str(e)}")
    return {
        "message": f"Database '{name}' loaded into session",
        "database": result["database"],
        "tables": result["tables"],
        "table_count": result["table_count"],
        "relationships_count": result["relationships_count"],
    }


@router.delete("/{name}")
async def delete_database(name: str) -> Dict:
    if not db_manager.database_exists(name):
        raise HTTPException(status_code=404, detail=f"Database '{name}' not found")

    session_id = "default"
    if session.current_database.get(session_id) == name:
        session.data_store[session_id] = {}
        session.current_database[session_id] = ""
        session.sensitive_columns_store[session_id] = {}
        session.relationships_store[session_id] = []

    try:
        db_manager.delete_database(name)
    except DatabaseError as e:
        raise HTTPException(status_code=500, detail=str(e))

    return {"message": f"Database '{name}' deleted successfully"}


def _list_excel_sheets(file_path: str) -> list:
    """List sheet names in an XLSX/XLS file using openpyxl (no polars needed)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return list(names)
    except Exception:
        return []


def _read_excel_sheet(file_path: str, sheet: str) -> pl.DataFrame:
    """Read a single sheet from an XLSX/XLS file as a polars DataFrame."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            raise HTTPException(
                status_code=404,
                detail=f"Sheet '{sheet}' not found in '{os.path.basename(file_path)}'",
            )
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return pl.DataFrame()
        header = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(rows[0])]
        data = rows[1:]
        return pl.DataFrame(data, schema=header, orient="row")
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to read sheet: {exc}")


def _build_table_file_map(db_name: str) -> Dict[str, str]:
    """Map in-memory table name -> on-disk file path (CSV or Excel sheet)."""
    mapping: Dict[str, str] = {}
    file_paths = db_manager.list_files(db_name)
    for fp in file_paths:
        fname = os.path.basename(fp)
        lower = fname.lower()
        if lower.endswith(".csv"):
            mapping[fname] = fp
        elif lower.endswith((".xlsx", ".xls")):
            sheets = _list_excel_sheets(fp)
            for sheet_name in sheets:
                mapping[f"{fname}::{sheet_name}"] = fp
    return mapping


def _load_table_for_download(file_path: str, table_name: str) -> pl.DataFrame:
    lower = file_path.lower()
    if lower.endswith(".csv"):
        return pl.read_csv(file_path)
    if lower.endswith((".xlsx", ".xls")):
        if "::" in table_name:
            sheet = table_name.split("::", 1)[1]
        else:
            sheet = table_name
        return _read_excel_sheet(file_path, sheet)
    raise HTTPException(status_code=400, detail=f"Unsupported file type: {file_path}")


def _iter_file(path: str, chunk_size: int = 8192):
    with open(path, "rb") as f:
        while True:
            chunk = f.read(chunk_size)
            if not chunk:
                break
            yield chunk


@router.get("/{name}/tables/{table_name:path}/download")
async def download_table(name: str, table_name: str, format: str = "csv") -> StreamingResponse:
    if format not in ("csv", "xlsx"):
        raise HTTPException(
            status_code=400,
            detail=f"format must be 'csv' or 'xlsx' (got '{format}')",
        )
    if not db_manager.database_exists(name):
        raise HTTPException(status_code=404, detail=f"Database '{name}' not found")

    file_map = _build_table_file_map(name)
    file_path: Optional[str] = file_map.get(table_name)
    if not file_path:
        raise HTTPException(
            status_code=404,
            detail=f"Table '{table_name}' not found in database '{name}'",
        )

    if format == "csv":
        if not file_path.lower().endswith(".csv"):
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Table '{table_name}' is backed by an Excel file. "
                    "Re-upload as CSV to enable direct download, or use format=xlsx."
                ),
            )
        try:
            df = pl.read_csv(file_path)
        except Exception as exc:
            raise HTTPException(status_code=500, detail=f"Failed to read table: {exc}")
        download_name = os.path.basename(table_name)
        return StreamingResponse(
            _iter_file(file_path),
            media_type="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": f'attachment; filename="{download_name}"',
                "X-Rows": str(df.height),
                "X-Columns": str(df.width),
            },
        )

    df = _load_table_for_download(file_path, table_name)
    buf = BytesIO()
    try:
        df.write_excel(buf)
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Failed to render XLSX: {exc}",
        )
    buf.seek(0)
    download_name = os.path.basename(table_name).replace(".csv", ".xlsx")
    if not download_name.endswith(".xlsx"):
        download_name = download_name + ".xlsx"
    return StreamingResponse(
        iter(lambda: buf.read(8192), b""),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{download_name}"',
            "X-Rows": str(df.height),
            "X-Columns": str(df.width),
        },
    )
