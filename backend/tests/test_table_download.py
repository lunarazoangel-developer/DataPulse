"""Verify the table download endpoint serves CSV and XLSX correctly."""
import io
import os
import tempfile
from pathlib import Path

import polars as pl
import pytest
from fastapi.testclient import TestClient

from api.routes import databases as db_routes


@pytest.fixture
def fake_db(tmp_path, monkeypatch):
    """Create a fake saved database with one CSV on disk and a stubbed manager."""
    db_dir = tmp_path / "2026-01-01_00-00-00"
    db_dir.mkdir()
    csv_path = db_dir / "users.csv"
    pl.DataFrame({"id": [1, 2, 3], "name": ["a", "b", "c"]}).write_csv(str(csv_path))
    meta_path = db_dir / "meta.json"
    meta_path.write_text(
        '{"name": "2026-01-01_00-00-00", "created_at": "2026-01-01T00:00:00", '
        '"tables": [{"name": "users.csv", "rows": 3, "columns": 2}], '
        '"file_count": 1, "total_rows": 3}'
    )

    monkeypatch.setattr(db_routes.db_manager, "database_exists", lambda n: True)
    monkeypatch.setattr(db_routes.db_manager, "list_files", lambda n: [str(csv_path)])
    monkeypatch.setattr(
        db_routes.db_manager,
        "list_databases",
        lambda: [],
    )
    return csv_path


def test_download_csv_returns_correct_content(fake_db):
    from main import app
    client = TestClient(app)
    response = client.get("/api/databases/test/tables/users.csv/download?format=csv")
    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "attachment" in response.headers["content-disposition"]
    assert response.headers["x-rows"] == "3"
    body = response.content.decode("utf-8")
    assert "id,name" in body
    assert "1,a" in body
    assert "3,c" in body


def test_download_xlsx_returns_excel_blob(fake_db):
    from main import app
    client = TestClient(app)
    response = client.get("/api/databases/test/tables/users.csv/download?format=xlsx")
    assert response.status_code == 200
    media = response.headers["content-type"]
    assert "spreadsheetml" in media or "xlsx" in media
    body = response.content
    assert len(body) > 0
    assert body[:2] == b"PK"  # XLSX is a ZIP file
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(body))
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("id", "name")
    assert (1, "a") in rows


def test_download_invalid_format_returns_400(fake_db):
    from main import app
    client = TestClient(app)
    response = client.get("/api/databases/test/tables/users.csv/download?format=pdf")
    assert response.status_code == 400
    assert "csv" in response.json()["detail"].lower() and "xlsx" in response.json()["detail"].lower()


def test_download_missing_table_returns_404(fake_db):
    from main import app
    client = TestClient(app)
    response = client.get("/api/databases/test/tables/missing.csv/download?format=csv")
    assert response.status_code == 404


def test_download_missing_database_returns_404(fake_db, monkeypatch):
    from main import app
    client = TestClient(app)
    monkeypatch.setattr(db_routes.db_manager, "database_exists", lambda n: False)
    response = client.get("/api/databases/nope/tables/users.csv/download?format=csv")
    assert response.status_code == 404


def test_download_csv_from_excel_backed_table_returns_helpful_error(tmp_path, monkeypatch):
    """If the table is backed by .xlsx, format=csv must give a clear message."""
    import openpyxl
    db_dir = tmp_path / "2026-02-02_00-00-00"
    db_dir.mkdir()
    xlsx_path = db_dir / "data.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["id", "value"])
    ws.append([1, "x"])
    wb.save(str(xlsx_path))

    monkeypatch.setattr(db_routes.db_manager, "database_exists", lambda n: True)
    monkeypatch.setattr(db_routes.db_manager, "list_files", lambda n: [str(xlsx_path)])

    from main import app
    client = TestClient(app)
    response = client.get("/api/databases/test/tables/data.xlsx::Sheet1/download?format=csv")
    assert response.status_code == 400
    assert "xlsx" in response.json()["detail"].lower()
